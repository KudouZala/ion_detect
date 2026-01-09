"""AutoEIS/
py文件，用于处理xlsx文件，我指定文件夹下的xlsx文件，我还要指定输出的xlsx文件的文件夹和文件名
以该xlsx文件的第二行的B列到K列的单元格内容作为title，第3行往后的B列-K列即为其数据，每列的数据要分组。

该数据所在行的A列的单元格内容进行判断，如果内容中含有ion_column且不含ion_column_renew，那么属于A组；
如果内容中含有ion且不含ion_column和ion_column_renew，那么属于B组；
如果内容中含有ion_column_renew，那么属于C组；

我需要你计算每个title下的每个组之间的平均差距，包括：B组的平均值-A组的平均值，C组的平均值-B组的平均值；C组的平均值-A组的平均值
在计算平均值的时候，你需要考虑误差值，去除那些明显存在测量错误的数据，
最后结果你要输出一个xlsx文件，该文件第一行是title，如：R1 P2w P2n...；第二行是B组的平均值-A组的平均值；第三和第四行是：C组的平均值-B组的平均值，C组的平均值-A组的平均值；
第一列是文字：B组的平均值-A组的平均值，C组的平均值-B组的平均值，C组的平均值-A组的平均值；
第一行第一列空出来

"""
import os
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import font_manager, rcParams
import time
# 设置字体路径
font_path = '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc'  # 替换为你需要的字体路径
font_prop = font_manager.FontProperties(fname=font_path)

# 设置全局字体
rcParams['font.family'] = font_prop.get_name()
rcParams['axes.unicode_minus'] = False  # 解决负号显示问题



def process_xlsx_data_analysis(input_folder, file_name, output_folder, output_filename_xlsx, output_filename_png, color_ion,custom_circuit):
    # 确保输出文件夹存在
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # 创建一个新的 workbook 用于输出结果
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active

    # 打开每个xlsx文件
    file_path = os.path.join(input_folder, file_name)
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    if custom_circuit == "R1-[P2,R3]-[P4,R5]-[P6,R7]":
        # 获取第二行B到K列作为title
        titles = [sheet.cell(row=2, column=i).value for i in range(2, 12)]
    elif custom_circuit == "R1-[P2,R3]-[P4,R5]":
        titles = [sheet.cell(row=2, column=i).value for i in range(2, 9)]
    
    # 获取从第三行开始的数据（B到K列）
    data = []
    for row in range(3, sheet.max_row + 1):
        if custom_circuit == "R1-[P2,R3]-[P4,R5]-[P6,R7]":
            data_row = [sheet.cell(row=row, column=i).value for i in range(2, 12)]
        elif custom_circuit == "R1-[P2,R3]-[P4,R5]":
            data_row = [sheet.cell(row=row, column=i).value for i in range(2, 9)]
        data.append(data_row)
    
    # 获取A列的内容
    a_column = [sheet.cell(row=row, column=1).value for row in range(3, sheet.max_row + 1)]
    
    # 判断分组
    group_a = {title: [] for title in titles}
    group_b = {title: [] for title in titles}
    group_c = {title: [] for title in titles}
    
    for i, row in enumerate(data):
        label = a_column[i]
        for col_idx, value in enumerate(row):
            title = titles[col_idx]
            if 'ion_column' in label and 'ion_column_renew' not in label:
                group_a[title].append(value)
            elif 'ion' in label and 'ion_column' not in label and 'ion_column_renew' not in label:
                group_b[title].append(value)
            elif 'ion_column_renew' in label:
                group_c[title].append(value)

    # 计算平均差距，去除误差
    def clean_and_average(group_data):
        """去除明显误差并计算均值，返回误差点的索引"""
        if len(group_data) < 2:
            return np.nan, []  # 如果数据太少，返回NaN和空列表
        
        # 计算中位数和MAD
        median = np.median(group_data)
        mad = np.median(np.abs(group_data - median))  # 中位数绝对偏差
        
        # 找到误差点的索引
        error_indices = [i for i, x in enumerate(group_data) if np.abs(x - median) > 3 * mad]
        
        # 清除误差点后的数据
        cleaned_data = [x for i, x in enumerate(group_data) if i not in error_indices]
        
        # 返回清理后的均值和误差点的索引
        return np.mean(cleaned_data) if cleaned_data else np.nan, error_indices


    # 准备输出的行数据
    row_a_b = []
    row_b_c = []
    row_c_a = []

    for title in titles:#每个title都计算每个组的平均值
        a_avg,a_error_index = clean_and_average(group_a[title])
        b_avg,b_error_index = clean_and_average(group_b[title])
        c_avg,c_error_index = clean_and_average(group_c[title])
        
        row_a_b.append(b_avg - a_avg if not np.isnan(b_avg) and not np.isnan(a_avg) else np.nan)
        row_b_c.append(c_avg - b_avg if not np.isnan(c_avg) and not np.isnan(b_avg) else np.nan)
        row_c_a.append(c_avg - a_avg if not np.isnan(c_avg) and not np.isnan(a_avg) else np.nan)

    # 将计算的结果写入输出文件
    # 第一行写入title
    ws_output.append([''] + titles)
    
    # 第二行：B组的平均值-A组的平均值
    ws_output.append(['B组的平均值-A组的平均值'] + row_a_b)
    
    # 第三行：C组的平均值-B组的平均值
    ws_output.append(['C组的平均值-B组的平均值'] + row_b_c)
    
    # 第四行：C组的平均值-A组的平均值
    ws_output.append(['C组的平均值-A组的平均值'] + row_c_a)

    # 保存输出的xlsx文件
    output_path = os.path.join(output_folder, output_filename_xlsx)
    wb_output.save(output_path)
    print(f"Results saved to {output_path}")

    # 绘图函数
    def plot_with_average_lines():
        # 创建一个 4 行 3 列的网格布局
        fig = plt.figure(figsize=(15, 20))
        if custom_circuit == "R1-[P2,R3]-[P4,R5]-[P6,R7]":
            grid_spec = fig.add_gridspec(4, 3)  # 4行3列的网格布局
        elif custom_circuit == "R1-[P2,R3]-[P4,R5]":
            grid_spec = fig.add_gridspec(3, 3)  # 4行3列的网格布局

        
        # 绘制第一行的图：只有一个图 R1
        ax1 = fig.add_subplot(grid_spec[0, 0])  # 第一行第一列
        col_idx = titles.index("R1")  # 找到"R1"在标题中的索引
        column_data = [data_row[col_idx] for data_row in data]

        

        
        # 分组数据
        group_a_data, group_b_data, group_c_data = [], [], []
        
        for j in range(len(data)):
            label = a_column[j]
            if 'ion_column' in label and 'ion_column_renew' not in label:
                group_a_data.append((j, column_data[j]))  # A组
            elif 'ion' in label and 'ion_column' not in label and 'ion_column_renew' not in label:
                group_b_data.append((j, column_data[j]))  # B组
            elif 'ion_column_renew' in label:
                group_c_data.append((j, column_data[j]))  # C组
        
        if group_b_data and group_c_data:
            for group_data, color, label in zip([group_a_data, group_b_data, group_c_data],
                                            [(192/255,192/255,192/255), color_ion, (160/255,160/255,160/255)],
                                            [f'R1_ion_column', f'R1_ion', f'R1_ion_column_renew']):
                group_indices, group_values = zip(*group_data)
                ax1.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2


        elif group_b_data:
            for group_data, color, label in zip([group_a_data, group_b_data],
                                            [(192/255,192/255,192/255), color_ion],
                                            [f'R1_ion_column', f'R1_ion']):
                group_indices, group_values = zip(*group_data)
                ax1.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2


        else:
            for group_data, color, label in zip([group_a_data],
                                            [(192/255,192/255,192/255)],
                                            [f'R1_ion_column']):
                group_indices, group_values = zip(*group_data)
                ax1.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2


        # 绘制不同组的数据点，分别指定颜色和标签
        
        # 绘制A组、B组、C组的平均值线
        if group_a_data:
            a_avg,_ = clean_and_average(group_a["R1"])
            if not np.isnan(a_avg):
                ax1.axhline(y=a_avg, color=(192/255,192/255,192/255), linestyle='--', label='ion_column_R1_average')
        if group_b_data:
            b_avg,_ = clean_and_average(group_b["R1"])
            if not np.isnan(b_avg):
                ax1.axhline(y=b_avg, color=color_ion, linestyle='--', label='ion_R1_average')
        if group_c_data:
            c_avg,_ = clean_and_average(group_c["R1"])
            if not np.isnan(c_avg):
                ax1.axhline(y=c_avg, color=(160/255,160/255,160/255), linestyle='--', label='ion_column_renew_R1_average')

        ax1.set_title("R1")
        ax1.set_xlabel('time/h')
        ax1.set_ylabel('value')
        ax1.legend()

        # 第二行依次放 R3, P2w, P2n
        second_row_titles = ["R3", "P2w", "P2n"]
        for i, title in enumerate(second_row_titles):
            ax = fig.add_subplot(grid_spec[1, i])  # 第二行的图
            col_idx = titles.index(title)  # 找到标题的索引
            column_data = [data_row[col_idx] for data_row in data]
            
            # 分组数据
            group_a_data, group_b_data, group_c_data = [], [], []
            for j in range(len(data)):
                label = a_column[j]
                if 'ion_column' in label and 'ion_column_renew' not in label:
                    group_a_data.append((j, column_data[j]))  # A组
                elif 'ion' in label and 'ion_column' not in label and 'ion_column_renew' not in label:
                    group_b_data.append((j, column_data[j]))  # B组
                elif 'ion_column_renew' in label:
                    group_c_data.append((j, column_data[j]))  # C组

            if group_b_data and group_c_data:
                    # 绘制不同组的数据点，分别指定颜色和标签
                for group_data, color, label in zip([group_a_data, group_b_data, group_c_data],
                                                        [(192/255,192/255,192/255), color_ion, (160/255,160/255,160/255)],
                                                        [f'{title}_ion_column', f'{title}_ion', f'{title}_ion_column_renew']):
                    group_indices, group_values = zip(*group_data)
                    ax.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2
            elif group_b_data:
                for group_data, color, label in zip([group_a_data, group_b_data],
                                                    [(192/255,192/255,192/255), color_ion],
                                                    [f'{title}_ion_column', f'{title}_ion']):
                    group_indices, group_values = zip(*group_data)
                    ax.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2
            else:
                for group_data, color, label in zip([group_a_data],
                                                    [(192/255,192/255,192/255)],
                                                    [f'{title}_ion_column']):
                    group_indices, group_values = zip(*group_data)
                    ax.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2
            
            # 绘制A组、B组、C组的平均值线
            # 绘制A组、B组、C组的平均值线
            a_avg,_ = clean_and_average(group_a[title])
            if not np.isnan(a_avg):
                ax.axhline(y=a_avg, color=(192/255,192/255,192/255), linestyle='--', label=f'{title}_ion_column_average')
            
            if group_b_data:

                b_avg,_ = clean_and_average(group_b[title])
                if not np.isnan(b_avg):
                    ax.axhline(y=b_avg, color=color_ion, linestyle='--', label=f'{title}_ion_average')

            if group_c_data:
                c_avg,_ = clean_and_average(group_c[title])
                if not np.isnan(c_avg):
                    ax.axhline(y=c_avg, color=(160/255,160/255,160/255), linestyle='--', label=f'{title}_ion_column_renew_H2SO4_average')

            ax.set_title(title)
            ax.set_xlabel('time/h')
            ax.set_ylabel('value')
            ax.legend()

        # 第三行依次放 R5, P4w, P4n
        third_row_titles = ["R5", "P4w", "P4n"]
        for i, title in enumerate(third_row_titles):
            ax = fig.add_subplot(grid_spec[2, i])  # 第三行的图
            col_idx = titles.index(title)  # 找到标题的索引
            column_data = [data_row[col_idx] for data_row in data]
            
            # 分组数据
            group_a_data, group_b_data, group_c_data = [], [], []
            for j in range(len(data)):
                label = a_column[j]
                if 'ion_column' in label and 'ion_column_renew' not in label:
                    group_a_data.append((j, column_data[j]))  # A组
                elif 'ion' in label and 'ion_column' not in label and 'ion_column_renew' not in label:
                    group_b_data.append((j, column_data[j]))  # B组
                elif 'ion_column_renew' in label:
                    group_c_data.append((j, column_data[j]))  # C组

            if group_b_data and group_c_data:
                    # 绘制不同组的数据点，分别指定颜色和标签
                for group_data, color, label in zip([group_a_data, group_b_data, group_c_data],
                                                        [(192/255,192/255,192/255), color_ion, (160/255,160/255,160/255)],
                                                        [f'{title}_ion_column', f'{title}_ion', f'{title}_ion_column_renew']):
                    group_indices, group_values = zip(*group_data)
                    ax.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2
            elif group_b_data:
                for group_data, color, label in zip([group_a_data, group_b_data],
                                                    [(192/255,192/255,192/255), color_ion],
                                                    [f'{title}_ion_column', f'{title}_ion']):
                    group_indices, group_values = zip(*group_data)
                    ax.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2
            else:
                for group_data, color, label in zip([group_a_data],
                                                    [(192/255,192/255,192/255)],
                                                    [f'{title}_ion_column']):
                    group_indices, group_values = zip(*group_data)
                    ax.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2
            
            # 绘制A组、B组、C组的平均值线
            # 绘制A组、B组、C组的平均值线
            a_avg,_ = clean_and_average(group_a[title])
            if not np.isnan(a_avg):
                ax.axhline(y=a_avg, color=(192/255,192/255,192/255), linestyle='--', label=f'{title}_ion_column_average')
            
            if group_b_data:

                b_avg,_ = clean_and_average(group_b[title])
                if not np.isnan(b_avg):
                    ax.axhline(y=b_avg, color=color_ion, linestyle='--', label=f'{title}_ion_average')

            if group_c_data:
                c_avg,_ = clean_and_average(group_c[title])
                if not np.isnan(c_avg):
                    ax.axhline(y=c_avg, color=(160/255,160/255,160/255), linestyle='--', label=f'{title}_ion_column_renew_H2SO4_average')

            ax.set_title(title)
            ax.set_xlabel('time/h')
            ax.set_ylabel('value')
            ax.legend()

        if custom_circuit == "R1-[P2,R3]-[P4,R5]-[P6,R7]":
            # 第四行依次放 R7, P6w, P6n
            fourth_row_titles = ["R7", "P6w", "P6n"]
            for i, title in enumerate(fourth_row_titles):
                ax = fig.add_subplot(grid_spec[3, i])  # 第四行的图
                col_idx = titles.index(title)  # 找到标题的索引
                column_data = [data_row[col_idx] for data_row in data]
                
                # 分组数据
                group_a_data, group_b_data, group_c_data = [], [], []
                for j in range(len(data)):
                    label = a_column[j]
                    if 'ion_column' in label and 'ion_column_renew' not in label:
                        group_a_data.append((j, column_data[j]))  # A组
                    elif 'ion' in label and 'ion_column' not in label and 'ion_column_renew' not in label:
                        group_b_data.append((j, column_data[j]))  # B组
                    elif 'ion_column_renew' in label:
                        group_c_data.append((j, column_data[j]))  # C组
                
                

                if group_b_data and group_c_data:
                    # 绘制不同组的数据点，分别指定颜色和标签
                    for group_data, color, label in zip([group_a_data, group_b_data, group_c_data],
                                                        [(192/255,192/255,192/255), color_ion, (160/255,160/255,160/255)],
                                                        [f'{title}_ion_column', f'{title}_ion', f'{title}_ion_column_renew']):
                        group_indices, group_values = zip(*group_data)
                        ax.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2
                elif group_b_data:
                    for group_data, color, label in zip([group_a_data, group_b_data],
                                                        [(192/255,192/255,192/255), color_ion],
                                                        [f'{title}_ion_column', f'{title}_ion']):
                        group_indices, group_values = zip(*group_data)
                        ax.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2
                else:
                    for group_data, color, label in zip([group_a_data],
                                                        [(192/255,192/255,192/255)],
                                                        [f'{title}_ion_column']):
                        group_indices, group_values = zip(*group_data)
                        ax.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2


                # 绘制A组、B组、C组的平均值线
                a_avg,_ = clean_and_average(group_a[title])
                if not np.isnan(a_avg):
                    ax.axhline(y=a_avg, color=(192/255,192/255,192/255), linestyle='--', label=f'{title}_ion_column_average')
                
                if group_b_data:

                    b_avg,_ = clean_and_average(group_b[title])
                    if not np.isnan(b_avg):
                        ax.axhline(y=b_avg, color=color_ion, linestyle='--', label=f'{title}_ion_average')

                if group_c_data:
                    c_avg,_ = clean_and_average(group_c[title])
                    if not np.isnan(c_avg):
                        ax.axhline(y=c_avg, color=(160/255,160/255,160/255), linestyle='--', label=f'{title}_ion_column_renew_H2SO4_average')

                ax.set_title(title)
                ax.set_xlabel('time/h')
                ax.set_ylabel('value')
                ax.legend()
        
        # 保存matplotlib图像为PNG
        output_path_png = os.path.join(output_folder, output_filename_png)
        plt.tight_layout()
        plt.savefig(output_path_png, dpi=300)
        plt.close()
        print(f"Plot saved to {output_path_png}")
        time.sleep(2)
    
    def plot_with_average_lines_clean():
        # 创建一个 4 行 3 列的网格布局
        fig = plt.figure(figsize=(15, 20))
        if custom_circuit == "R1-[P2,R3]-[P4,R5]-[P6,R7]":
            grid_spec = fig.add_gridspec(4, 3)  # 4行3列的网格布局
        elif custom_circuit == "R1-[P2,R3]-[P4,R5]":
            grid_spec = fig.add_gridspec(3, 3)  # 4行3列的网格布局

        
        # 绘制第一行的图：只有一个图 R1
        ax1 = fig.add_subplot(grid_spec[0, 0])  # 第一行第一列
        col_idx = titles.index("R1")  # 找到"R1"在标题中的索引
        column_data = [data_row[col_idx] for data_row in data]

        # 绘制A组、B组、C组的平均值线
        if group_a["R1"]:
            a_avg,a_error_index = clean_and_average(group_a["R1"])
            if not np.isnan(a_avg):
                ax1.axhline(y=a_avg, color=(192/255,192/255,192/255), linestyle='--', label='ion_column_R1_average')
        if group_b["R1"]:
            b_avg,b_error_index = clean_and_average(group_b["R1"])
            b_error_index = [index + len(group_a["R1"]) for index in b_error_index]
            if not np.isnan(b_avg):
                ax1.axhline(y=b_avg, color=color_ion, linestyle='--', label='ion_R1_average')
        if group_c["R1"]:
            c_avg,c_error_index = clean_and_average(group_c["R1"])
            c_error_index = [index + len(group_a["R1"])+len(group_b["R1"]) for index in c_error_index]
            if not np.isnan(c_avg):
                ax1.axhline(y=c_avg, color=(160/255,160/255,160/255), linestyle='--', label='ion_column_renew_R1_average')
        

        
        # 分组数据
        group_a_data, group_b_data, group_c_data = [], [], []
        
        for j in range(len(data)):
            label = a_column[j]
            
            if 'ion_column' in label and 'ion_column_renew' not in label:
                if j in a_error_index:
                    print(f"R1跳过第{j}个点")
                    continue
                group_a_data.append((j, column_data[j]))  # A组
            elif 'ion' in label and 'ion_column' not in label and 'ion_column_renew' not in label:
                if j in b_error_index:
                    print(f"R1跳过第{j}个点")
                    continue
                group_b_data.append((j, column_data[j]))  # B组
            elif 'ion_column_renew' in label:
                if j in c_error_index:
                    print(f"R1跳过第{j}个点")
                    continue
                group_c_data.append((j, column_data[j]))  # C组
        
        if group_b_data and group_c_data:
            for group_data, color, label in zip([group_a_data, group_b_data, group_c_data],
                                            [(192/255,192/255,192/255), color_ion, (160/255,160/255,160/255)],
                                            [f'R1_ion_column', f'R1_ion', f'R1_ion_column_renew']):
                group_indices, group_values = zip(*group_data)
                ax1.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2


        elif group_b_data:
            for group_data, color, label in zip([group_a_data, group_b_data],
                                            [(192/255,192/255,192/255), color_ion],
                                            [f'R1_ion_column', f'R1_ion']):
                group_indices, group_values = zip(*group_data)
                ax1.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2


        else:
            for group_data, color, label in zip([group_a_data],
                                            [(192/255,192/255,192/255)],
                                            [f'R1_ion_column']):
                group_indices, group_values = zip(*group_data)
                ax1.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2


        # 绘制不同组的数据点，分别指定颜色和标签
        
        ax1.set_title("R1")
        ax1.set_xlabel('time/h')
        ax1.set_ylabel('value')
        ax1.legend()

        # 第二行依次放 R3, P2w, P2n
        second_row_titles = ["R3", "P2w", "P2n"]
        for i, title in enumerate(second_row_titles):
            ax = fig.add_subplot(grid_spec[1, i])  # 第二行的图
            col_idx = titles.index(title)  # 找到标题的索引
            column_data = [data_row[col_idx] for data_row in data]

            # 绘制A组、B组、C组的平均值线
            # 绘制A组、B组、C组的平均值线

            a_avg,a_error_index = clean_and_average(group_a[title])
            if not np.isnan(a_avg):
                ax.axhline(y=a_avg, color=(192/255,192/255,192/255), linestyle='--', label=f'{title}_ion_column_average')
            
            if group_b[title]:
                b_avg,b_error_index = clean_and_average(group_b[title])
                b_error_index = [index + len(group_a[title]) for index in b_error_index]
                if not np.isnan(b_avg):
                    ax.axhline(y=b_avg, color=color_ion, linestyle='--', label=f'{title}_ion_average')

            if group_c[title]:
                c_avg,c_error_index = clean_and_average(group_c[title])
                c_error_index = [index + len(group_a[title])+len(group_b[title]) for index in c_error_index]
                if not np.isnan(c_avg):
                    ax.axhline(y=c_avg, color=(160/255,160/255,160/255), linestyle='--', label=f'{title}_ion_column_renew_H2SO4_average')

            
            # 分组数据
            group_a_data, group_b_data, group_c_data = [], [], []
            for j in range(len(data)):
                
                label = a_column[j]
                if 'ion_column' in label and 'ion_column_renew' not in label:
                    if j in a_error_index:
                        print(f"{title}跳过第{j}个点")
                        continue  # 跳过误差点
                    group_a_data.append((j, column_data[j]))  # A组
                elif 'ion' in label and 'ion_column' not in label and 'ion_column_renew' not in label:
                    if j in b_error_index:
                        print(f"{title}跳过第{j}个点")
                        continue  # 跳过误差点
                    group_b_data.append((j, column_data[j]))  # B组
                elif 'ion_column_renew' in label:
                    if j in c_error_index:
                        print(f"{title}跳过第{j}个点")
                        continue  # 跳过误差点
                    group_c_data.append((j, column_data[j]))  # C组

            if group_b_data and group_c_data:
                    # 绘制不同组的数据点，分别指定颜色和标签
                for group_data, color, label in zip([group_a_data, group_b_data, group_c_data],
                                                        [(192/255,192/255,192/255), color_ion, (160/255,160/255,160/255)],
                                                        [f'{title}_ion_column', f'{title}_ion', f'{title}_ion_column_renew']):
                    group_indices, group_values = zip(*group_data)
                    ax.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2
            elif group_b_data:
                for group_data, color, label in zip([group_a_data, group_b_data],
                                                    [(192/255,192/255,192/255), color_ion],
                                                    [f'{title}_ion_column', f'{title}_ion']):
                    group_indices, group_values = zip(*group_data)
                    ax.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2
            else:
                for group_data, color, label in zip([group_a_data],
                                                    [(192/255,192/255,192/255)],
                                                    [f'{title}_ion_column']):
                    group_indices, group_values = zip(*group_data)
                    ax.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2
            
            

            ax.set_title(title)
            ax.set_xlabel('time/h')
            ax.set_ylabel('value')
            ax.legend()

        # 第三行依次放 R5, P4w, P4n
        third_row_titles = ["R5", "P4w", "P4n"]
        for i, title in enumerate(third_row_titles):
            ax = fig.add_subplot(grid_spec[2, i])  # 第三行的图
            col_idx = titles.index(title)  # 找到标题的索引
            column_data = [data_row[col_idx] for data_row in data]

            a_avg,a_error_index = clean_and_average(group_a[title])
            if not np.isnan(a_avg):
                ax.axhline(y=a_avg, color=(192/255,192/255,192/255), linestyle='--', label=f'{title}_ion_column_average')
            
            if group_b[title]:

                b_avg,b_error_index = clean_and_average(group_b[title])
                b_error_index = [index + len(group_a[title]) for index in b_error_index]
                if not np.isnan(b_avg):
                    ax.axhline(y=b_avg, color=color_ion, linestyle='--', label=f'{title}_ion_average')

            if group_c[title]:
                c_avg,c_error_index = clean_and_average(group_c[title])
                c_error_index = [index + len(group_a[title])+len(group_b[title]) for index in c_error_index]
                if not np.isnan(c_avg):
                    ax.axhline(y=c_avg, color=(160/255,160/255,160/255), linestyle='--', label=f'{title}_ion_column_renew_H2SO4_average')
            
            # 分组数据
            group_a_data, group_b_data, group_c_data = [], [], []
            for j in range(len(data)):
                
                label = a_column[j]
                if 'ion_column' in label and 'ion_column_renew' not in label:
                    if j in a_error_index:
                        print(f"{title}跳过第{j}个点")
                        continue  # 跳过误差点
                    group_a_data.append((j, column_data[j]))  # A组
                elif 'ion' in label and 'ion_column' not in label and 'ion_column_renew' not in label:
                    if j in b_error_index:
                        print(f"{title}跳过第{j}个点")
                        continue  # 跳过误差点
                    group_b_data.append((j, column_data[j]))  # B组
                elif 'ion_column_renew' in label:
                    if j in c_error_index:
                        print(f"{title}跳过第{j}个点")
                        continue  # 跳过误差点
                    group_c_data.append((j, column_data[j]))  # C组

            if group_b_data and group_c_data:
                    # 绘制不同组的数据点，分别指定颜色和标签
                for group_data, color, label in zip([group_a_data, group_b_data, group_c_data],
                                                        [(192/255,192/255,192/255), color_ion, (160/255,160/255,160/255)],
                                                        [f'{title}_ion_column', f'{title}_ion', f'{title}_ion_column_renew']):
                    group_indices, group_values = zip(*group_data)
                    ax.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2
            elif group_b_data:
                for group_data, color, label in zip([group_a_data, group_b_data],
                                                    [(192/255,192/255,192/255), color_ion],
                                                    [f'{title}_ion_column', f'{title}_ion']):
                    group_indices, group_values = zip(*group_data)
                    ax.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2
            else:
                for group_data, color, label in zip([group_a_data],
                                                    [(192/255,192/255,192/255)],
                                                    [f'{title}_ion_column']):
                    group_indices, group_values = zip(*group_data)
                    ax.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2
            
            # 绘制A组、B组、C组的平均值线
            # 绘制A组、B组、C组的平均值线
            

            ax.set_title(title)
            ax.set_xlabel('time/h')
            ax.set_ylabel('value')
            ax.legend()

        if custom_circuit == "R1-[P2,R3]-[P4,R5]-[P6,R7]":
            # 第四行依次放 R7, P6w, P6n
            fourth_row_titles = ["R7", "P6w", "P6n"]
            for i, title in enumerate(fourth_row_titles):
                ax = fig.add_subplot(grid_spec[3, i])  # 第四行的图
                col_idx = titles.index(title)  # 找到标题的索引
                column_data = [data_row[col_idx] for data_row in data]
                # 绘制A组、B组、C组的平均值线
                a_avg,a_error_index = clean_and_average(group_a[title])
                if not np.isnan(a_avg):
                    ax.axhline(y=a_avg, color=(192/255,192/255,192/255), linestyle='--', label=f'{title}_ion_column_average')
                
                if group_b[title]:
                    b_avg,b_error_index = clean_and_average(group_b[title])
                    b_error_index = [index + len(group_a[title]) for index in b_error_index]
                    if not np.isnan(b_avg):
                        ax.axhline(y=b_avg, color=color_ion, linestyle='--', label=f'{title}_ion_average')

                if group_c[title]:
                    c_avg,c_error_index = clean_and_average(group_c[title])
                    c_error_index = [index + len(group_a[title])+len(group_b[title]) for index in c_error_index]
                    if not np.isnan(c_avg):
                        ax.axhline(y=c_avg, color=(160/255,160/255,160/255), linestyle='--', label=f'{title}_ion_column_renew_H2SO4_average')
                
                # 分组数据
                group_a_data, group_b_data, group_c_data = [], [], []
                for j in range(len(data)):
                    
                    label = a_column[j]
                    if 'ion_column' in label and 'ion_column_renew' not in label:
                        if j in a_error_index:
                            print(f"{title}跳过第{j}个点")
                            continue  # 跳过误差点
                        group_a_data.append((j, column_data[j]))  # A组
                    elif 'ion' in label and 'ion_column' not in label and 'ion_column_renew' not in label:
                        if j in b_error_index:
                            print(f"{title}跳过第{j}个点")
                            continue  # 跳过误差点
                        group_b_data.append((j, column_data[j]))  # B组
                    elif 'ion_column_renew' in label:
                        if j in c_error_index:
                            print(f"{title}跳过第{j}个点")
                            continue  # 跳过误差点
                        group_c_data.append((j, column_data[j]))  # C组
                
                

                if group_b_data and group_c_data:
                    # 绘制不同组的数据点，分别指定颜色和标签
                    for group_data, color, label in zip([group_a_data, group_b_data, group_c_data],
                                                        [(192/255,192/255,192/255), color_ion, (160/255,160/255,160/255)],
                                                        [f'{title}_ion_column', f'{title}_ion', f'{title}_ion_column_renew']):
                        group_indices, group_values = zip(*group_data)
                        ax.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2
                elif group_b_data:
                    for group_data, color, label in zip([group_a_data, group_b_data],
                                                        [(192/255,192/255,192/255), color_ion],
                                                        [f'{title}_ion_column', f'{title}_ion']):
                        group_indices, group_values = zip(*group_data)
                        ax.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2
                else:
                    for group_data, color, label in zip([group_a_data],
                                                        [(192/255,192/255,192/255)],
                                                        [f'{title}_ion_column']):
                        group_indices, group_values = zip(*group_data)
                        ax.scatter([i * 2 for i in group_indices], group_values, c=color, label=label, marker='o')  # 横坐标*2

                ax.set_title(title)
                ax.set_xlabel('time/h')
                ax.set_ylabel('value')
                ax.legend()
        
        # 保存matplotlib图像为PNG
        output_filename_png_clean =  os.path.splitext(output_filename_png)[0] + '_clean' + os.path.splitext(output_filename_png)[1]
        output_path_png = os.path.join(output_folder, output_filename_png_clean)
        plt.tight_layout()
        plt.savefig(output_path_png, dpi=300)
        plt.close()
        print(f"Plot saved to {output_path_png}")
        time.sleep(2)

    

    # 调用绘图函数
    plot_with_average_lines()
    time.sleep(2)
    print("执行完毕:plot_with_average_lines")
    plot_with_average_lines_clean()
    time.sleep(2)
    print("执行完毕:plot_with_average_lines_clean")

# # # 调用函数
# input_folder = '/home/cagalii/Application/autoeis/AutoEIS/examples/eis_results/20241101_2ppm钙离子污染和恢复测试/output_ecm_firecloud'  # 输入文件夹路径
# input_filename = "output_values.xlsx"
# custom_circuit = "R1-[P2,R3]-[P4,R5]-[P6,R7]"

# output_folder = '/home/cagalii/Application/autoeis/AutoEIS/examples/eis_results/'  # 输出文件夹路径
# output_filename_xlsx = 'output_results_test.xlsx'  # 输出文件名
# output_filename_png = 'output_results_test.png'  # 输出文件名
# color_ion = (204/255, 255/255, 255/255)  # 传入 RGB 值，Matplotlib 需要将其归一化到 [0, 1]

# input_folder = '/home/cagalii/Application/autoeis/AutoEIS/examples/eis_results/20241101_2ppm铜离子污染和恢复测试_R1-[P2,R3]-[P4,R5]/output_ecm_firecloud'
# input_filename = 'output_values_R1-[P2,R3]-[P4,R5].xlsx'
# custom_circuit = "R1-[P2,R3]-[P4,R5]"


# input_folder = '/home/cagalii/Application/autoeis/AutoEIS/examples/eis_results/20241010_2ppm钠离子污染测试/output_ecm_firecloud'
# input_filename = 'output_values.xlsx'
# custom_circuit = "R1-[P2,R3]-[P4,R5]-[P6,R7]"
# process_xlsx_data_analysis(input_folder, input_filename, output_folder, output_filename_xlsx, output_filename_png, color_ion,custom_circuit)

